"""
Sistema de Asistencia con Reconocimiento Facial
Backend: FastAPI + face_recognition + OpenPyXL
"""

import os
import io
import base64
import json
import shutil
from datetime import datetime, date
from pathlib import Path
from typing import Optional

import face_recognition
import numpy as np
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

# ─── Configuración ─────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
STUDENTS_DIR = BASE_DIR / "students"
ATTENDANCE_DIR = BASE_DIR / "attendance"
ENCODINGS_FILE = BASE_DIR / "encodings.json"

STUDENTS_DIR.mkdir(exist_ok=True)
ATTENDANCE_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Sistema de Asistencia Facial", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Modelos ───────────────────────────────────────────────────────────────────
class StudentInfo(BaseModel):
    id: str
    name: str
    group: Optional[str] = ""

# ─── Utilidades ───────────────────────────────────────────────────────────────
def load_encodings() -> dict:
    """Carga los encodings guardados de alumnos."""
    if ENCODINGS_FILE.exists():
        with open(ENCODINGS_FILE, "r") as f:
            data = json.load(f)
        # Convertir listas a numpy arrays
        for student_id in data:
            data[student_id]["encodings"] = [
                np.array(enc) for enc in data[student_id]["encodings"]
            ]
        return data
    return {}

def save_encodings(encodings: dict):
    """Guarda los encodings en disco."""
    serializable = {}
    for student_id, info in encodings.items():
        serializable[student_id] = {
            "name": info["name"],
            "group": info.get("group", ""),
            "encodings": [enc.tolist() for enc in info["encodings"]],
        }
    with open(ENCODINGS_FILE, "w") as f:
        json.dump(serializable, f, indent=2, ensure_ascii=False)

def get_attendance_file(group: str = "General") -> Path:
    """Retorna la ruta del archivo Excel de asistencia del día."""
    today = date.today().strftime("%Y-%m-%d")
    safe_group = group.replace(" ", "_").replace("/", "-")
    return ATTENDANCE_DIR / f"Asistencia_{safe_group}_{today}.xlsx"

def create_or_load_workbook(path: Path, group: str) -> tuple:
    """Crea o carga un workbook de asistencia."""
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        return wb, ws

    wb = openpyxl.Workbook()
    ws = wb.active
    today = date.today().strftime("%d/%m/%Y")
    ws.title = "Asistencia"

    # ── Estilos ──────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    title_fill  = PatternFill("solid", fgColor="2E6DA4")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    title_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Fila 1: Título
    ws.merge_cells("A1:E1")
    ws["A1"] = f"📋 Lista de Asistencia — {group} — {today}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    # Fila 2: Encabezados
    headers = ["#", "ID Alumno", "Nombre", "Grupo", "Hora de Registro"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Anchos de columna
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.row_dimensions[2].height = 20

    wb.save(path)
    return wb, ws

def add_attendance_row(path: Path, student_id: str, name: str, group: str, time_str: str):
    """Agrega una fila de asistencia al Excel."""
    wb, ws = create_or_load_workbook(path, group)

    # Buscar si ya está registrado hoy
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1] == student_id:
            return False  # Ya registrado

    row_num = ws.max_row + 1
    num = row_num - 2

    row_fill = PatternFill("solid", fgColor="EBF3FB" if num % 2 == 0 else "FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    green_font = Font(color="1A7F3C", bold=True, name="Calibri")

    values = [num, student_id, name, group, time_str]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = row_fill
        cell.alignment = center
        cell.border = thin_border
        if col == 5:
            cell.font = green_font

    wb.save(path)
    return True

# ─── Endpoints ────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"message": "Sistema de Asistencia Facial activo 🎓", "version": "1.0.0"}

@app.get("/students")
def list_students():
    """Lista todos los alumnos registrados."""
    encodings = load_encodings()
    return [
        {"id": sid, "name": info["name"], "group": info.get("group", "")}
        for sid, info in encodings.items()
    ]

@app.post("/students/register")
async def register_student(
    name: str = Form(...),
    student_id: str = Form(...),
    group: str = Form("General"),
    photo: UploadFile = File(...),
):
    """Registra un alumno con su foto."""
    contents = await photo.read()
    img = face_recognition.load_image_file(io.BytesIO(contents))
    encs = face_recognition.face_encodings(img)

    if len(encs) == 0:
        raise HTTPException(400, "No se detectó ninguna cara en la foto.")
    if len(encs) > 1:
        raise HTTPException(400, "Se detectaron varias caras. Usa una foto individual.")

    # Guardar foto
    student_dir = STUDENTS_DIR / student_id
    student_dir.mkdir(exist_ok=True)
    photo_path = student_dir / photo.filename
    with open(photo_path, "wb") as f:
        f.write(contents)

    # Guardar encoding
    encodings = load_encodings()
    if student_id not in encodings:
        encodings[student_id] = {"name": name, "group": group, "encodings": []}
    encodings[student_id]["encodings"].append(encs[0])
    encodings[student_id]["name"] = name
    encodings[student_id]["group"] = group
    save_encodings(encodings)

    return {"success": True, "message": f"Alumno '{name}' registrado correctamente."}

@app.delete("/students/{student_id}")
def delete_student(student_id: str):
    """Elimina un alumno del sistema."""
    encodings = load_encodings()
    if student_id not in encodings:
        raise HTTPException(404, "Alumno no encontrado.")
    
    name = encodings[student_id]["name"]
    del encodings[student_id]
    save_encodings(encodings)

    student_dir = STUDENTS_DIR / student_id
    if student_dir.exists():
        shutil.rmtree(student_dir)

    return {"success": True, "message": f"Alumno '{name}' eliminado."}

@app.post("/attendance/recognize")
async def recognize_attendance(
    photo: UploadFile = File(...),
    group: str = Form("General"),
    tolerance: float = Form(0.5),
):
    """
    Recibe una foto de grupo, detecta caras, las compara con la base de datos
    y registra la asistencia de los alumnos reconocidos.
    """
    encodings_db = load_encodings()
    if not encodings_db:
        raise HTTPException(400, "No hay alumnos registrados aún.")

    # Cargar imagen recibida
    contents = await photo.read()
    img = face_recognition.load_image_file(io.BytesIO(contents))

    # Detectar caras en la foto
    face_locations = face_recognition.face_locations(img, model="hog")
    if not face_locations:
        return {"recognized": [], "unknown": 0, "total_faces": 0}

    face_encs = face_recognition.face_encodings(img, face_locations)

    # Preparar encodings conocidos
    known_ids      = []
    known_names    = []
    known_groups   = []
    known_encs     = []

    for sid, info in encodings_db.items():
        for enc in info["encodings"]:
            known_ids.append(sid)
            known_names.append(info["name"])
            known_groups.append(info.get("group", group))
            known_encs.append(enc)

    recognized = []
    unknown_count = 0
    now = datetime.now().strftime("%H:%M:%S")
    attendance_path = get_attendance_file(group)
    create_or_load_workbook(attendance_path, group)  # asegura que existe

    seen_ids = set()  # evitar duplicados en la misma foto

    for face_enc in face_encs:
        distances = face_recognition.face_distance(known_encs, face_enc)
        best_idx = int(np.argmin(distances))
        best_dist = distances[best_idx]

        if best_dist <= tolerance:
            sid   = known_ids[best_idx]
            name  = known_names[best_idx]
            grp   = known_groups[best_idx]

            if sid not in seen_ids:
                seen_ids.add(sid)
                already = not add_attendance_row(attendance_path, sid, name, grp, now)
                recognized.append({
                    "id": sid,
                    "name": name,
                    "group": grp,
                    "confidence": round((1 - best_dist) * 100, 1),
                    "already_registered": already,
                    "time": now,
                })
        else:
            unknown_count += 1

    return {
        "recognized": recognized,
        "unknown": unknown_count,
        "total_faces": len(face_encs),
        "attendance_file": attendance_path.name,
    }

@app.get("/attendance/today")
def get_today_attendance(group: str = "General"):
    """Retorna la asistencia del día actual."""
    path = get_attendance_file(group)
    if not path.exists():
        return {"date": date.today().isoformat(), "students": [], "total": 0}

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    students = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1]:  # tiene ID
            students.append({
                "num": row[0],
                "id": row[1],
                "name": row[2],
                "group": row[3],
                "time": row[4],
            })

    return {"date": date.today().isoformat(), "students": students, "total": len(students)}

@app.get("/attendance/download")
def download_attendance(group: str = "General"):
    """Descarga el Excel de asistencia del día."""
    path = get_attendance_file(group)
    if not path.exists():
        raise HTTPException(404, "No hay asistencia registrada para hoy.")
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=path.name,
    )

@app.get("/groups")
def list_groups():
    """Lista los grupos disponibles."""
    encodings = load_encodings()
    groups = list(set(info.get("group", "General") for info in encodings.values()))
    return sorted(groups)
